import openpyxl
from django.shortcuts import render,redirect
from django.http import HttpResponse
import os
from django.conf import settings

def home(request):
    download_url = None
    extracted_url = None

    if request.method == 'POST' and 'excel_file1' in request.FILES and 'excel_file2' in request.FILES and 'excel_file3' in request.FILES:
        excel_file1 = request.FILES['excel_file1']
        excel_file2 = request.FILES['excel_file2']
        excel_file3 = request.FILES['excel_file3']

        wb1 = openpyxl.load_workbook(excel_file1)
        wb2 = openpyxl.load_workbook(excel_file2)
        wb3 = openpyxl.load_workbook(excel_file3)

        sheet1 = wb1.active
        sheet2 = wb2.active
        sheet3 = wb3.active

        register_map = {}
        for row in range(2, sheet2.max_row + 1):
            reg_no = sheet2.cell(row=row, column=2).value
            sgpa = sheet2.cell(row=row, column=4).value
            cgpa = sheet2.cell(row=row, column=5).value
            if reg_no:
                register_map[reg_no] = (sgpa, cgpa)

        next_col_sgpa = sheet1.max_column + 1
        next_col_cgpa = next_col_sgpa + 1

        for row in range(2, sheet1.max_row + 1):
            reg_no = sheet1.cell(row=row, column=2).value
            if reg_no and reg_no in register_map:
                sgpa, cgpa = register_map[reg_no]
                sheet1.cell(row=row, column=next_col_sgpa, value=sgpa)
                sheet1.cell(row=row, column=next_col_cgpa, value=cgpa)

        merged_file_path = os.path.join(settings.MEDIA_ROOT, 'merged.xlsx')
        wb1.save(merged_file_path)
        download_url = f'{settings.MEDIA_URL}merged.xlsx'

        merged_wb = openpyxl.load_workbook(merged_file_path)
        merged_sheet = merged_wb.active

        extracted_wb = openpyxl.Workbook()
        extracted_sheet = extracted_wb.active

        headers = ["S.No"] + [merged_sheet.cell(row=2, column=col).value for col in range(2, merged_sheet.max_column + 1)]
        for col, header in enumerate(headers, start=1):
            extracted_sheet.cell(row=1, column=col, value=header)

        row_counter = 2
        for row in range(2, sheet3.max_row + 1):
            reg_no = sheet3.cell(row=row, column=2).value
            for merged_row in range(2, merged_sheet.max_row + 1):
                merged_reg_no = merged_sheet.cell(row=merged_row, column=2).value
                if reg_no == merged_reg_no:
                    extracted_sheet.cell(row=row_counter, column=1, value=row_counter - 1)  # S.No
                    for col in range(2, merged_sheet.max_column + 1):
                        extracted_sheet.cell(row=row_counter, column=col, value=merged_sheet.cell(row=merged_row, column=col).value)
                    row_counter += 1
                    break

        extracted_file_path = os.path.join(settings.MEDIA_ROOT, 'extracted.xlsx')
        extracted_wb.save(extracted_file_path)
        extracted_url = f'{settings.MEDIA_URL}extracted.xlsx'
        return redirect('about')

    return render(request, 'home.html', {'download_url': download_url, 'extracted_url': extracted_url})
import pandas as pd
import json
def about(request):
    # Path to the extracted Excel file
    merged_file_path = os.path.join(settings.MEDIA_ROOT, 'extracted.xlsx')

    # Load the Excel file into a DataFrame
    df = pd.read_excel(merged_file_path)
    
    # Convert DataFrame to a list of dictionaries
    students = df.to_dict(orient='records')
    valid=make_valid_identifier(students)
    # Sort students by CGPA and SGPA
    sorted_students = sorted(valid, key=lambda x: (x['cgpa'], x['sgpa']), reverse=True)

    # Convert the sorted student data to JSON
    students_json = json.dumps(sorted_students)
    
    # Pass the JSON data to the template
    return render(request, 'about.html', {'students_json': students_json})

import re

def make_valid_identifier(data):
    # Function to convert dictionary keys to valid Python identifiers
    def convert_key(key):
        # Convert to lowercase and replace spaces with underscores
        key = key.strip().lower()
        # Remove any characters that are not alphanumeric or underscores
        key = re.sub(r'[^a-z0-9_]', '', key)
        return key
    
    # Iterate over the list of dictionaries and convert keys for each dictionary
    return [
        {convert_key(k): v for k, v in student.items()} 
        for student in data
    ]
